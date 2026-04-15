"""
Read gsc_full_export.json and export all tables to a single Excel workbook (.xlsx).
Each main data set becomes a separate sheet.
Requires: pip install openpyxl
"""
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
from gsc_paths import GSC_DATA_DIR
INPUT_JSON = os.path.join(GSC_DATA_DIR, "gsc_full_export.json")
OUTPUT_XLSX = os.path.join(GSC_DATA_DIR, "gsc_full_export.xlsx")


def sheet_from_list(ws, rows, sheet_title_max=31):
    """Write list of dicts to sheet. First row = headers."""
    if not rows:
        return
    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h)
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=val)


def flatten_sitemap(s):
    """Turn sitemap entry into flat dict for one row."""
    if isinstance(s, dict):
        out = {"path": s.get("path"), "lastSubmitted": s.get("lastSubmitted"), "lastDownloaded": s.get("lastDownloaded"),
               "isPending": s.get("isPending"), "isSitemapsIndex": s.get("isSitemapsIndex"), "warnings": s.get("warnings"), "errors": s.get("errors"), "type": s.get("type")}
        contents = s.get("contents") or []
        for i, c in enumerate(contents):
            if isinstance(c, dict):
                out[f"content_type_{i}"] = c.get("type")
                out[f"content_submitted_{i}"] = c.get("submitted")
                out[f"content_indexed_{i}"] = c.get("indexed")
        return out
    return {"path": str(s)}


def flatten_url_inspection(url, data):
    """One row per URL with key index fields."""
    if isinstance(data, dict) and "error" in data:
        return {"url": url, "error": data["error"]}
    out = {"url": url}
    idx = data.get("indexStatusResult") or {}
    out["indexingState"] = idx.get("indexingState")
    out["coverageState"] = idx.get("coverageState")
    out["lastCrawlTime"] = idx.get("lastCrawlTime")
    out["robotsTxtState"] = idx.get("robotsTxtState")
    out["verdict"] = idx.get("verdict")
    return out


def main():
    if not os.path.isfile(INPUT_JSON):
        print("Input not found:", INPUT_JSON)
        return
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Meta"
    meta = data.get("meta") or {}
    for i, (k, v) in enumerate(meta.items(), 1):
        ws0.cell(row=i, column=1, value=k)
        ws0.cell(row=i, column=2, value=str(v))

    # Sites
    sites = data.get("sites")
    if isinstance(sites, list) and sites:
        ws = wb.create_sheet("Sites")
        sheet_from_list(ws, sites)
    elif isinstance(sites, dict):
        ws = wb.create_sheet("Sites")
        ws.cell(row=1, column=1, value="error")
        ws.cell(row=1, column=2, value=sites.get("error", ""))

    # Sitemaps
    sitemaps = data.get("sitemaps")
    if isinstance(sitemaps, list) and sitemaps:
        ws = wb.create_sheet("Sitemaps")
        flat = [flatten_sitemap(s) for s in sitemaps]
        sheet_from_list(ws, flat)
    elif isinstance(sitemaps, dict):
        ws = wb.create_sheet("Sitemaps")
        ws.cell(row=1, column=1, value="error")
        ws.cell(row=1, column=2, value=sitemaps.get("error", ""))

    # Search Analytics sheets
    sa = data.get("searchAnalytics") or {}
    sheet_specs = [
        ("by_date", "Performance_by_Date"),
        ("by_query", "Queries"),
        ("by_page", "Pages"),
        ("by_country", "By_Country"),
        ("by_device", "By_Device"),
        ("by_date_and_device", "Date_Device"),
        ("by_date_and_country", "Date_Country"),
        ("by_query_and_page", "Query_Page"),
    ]
    for key, sheet_name in sheet_specs:
        rows = sa.get(key)
        if isinstance(rows, list) and rows:
            name = sheet_name[:31]
            ws = wb.create_sheet(name)
            sheet_from_list(ws, rows)

    # URL Inspection
    insp = data.get("urlInspection") or {}
    if insp and "_error" not in str(insp.get("_error", "")):
        rows = [flatten_url_inspection(url, insp[url]) for url in insp if not url.startswith("_")]
        if rows:
            ws = wb.create_sheet("URL_Inspection")
            sheet_from_list(ws, rows)
    elif insp.get("_error"):
        ws = wb.create_sheet("URL_Inspection")
        ws.cell(row=1, column=1, value="error")
        ws.cell(row=1, column=2, value=insp["_error"])

    os.makedirs(os.path.dirname(OUTPUT_XLSX) or ".", exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print("Saved:", os.path.abspath(OUTPUT_XLSX))
    print("Sheets:", [s.title for s in wb.worksheets])


if __name__ == "__main__":
    main()
